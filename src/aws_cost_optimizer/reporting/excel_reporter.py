"""
Excel Reporter module for generating comprehensive cost optimization reports
Creates formatted Excel workbooks with charts and summaries
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import logging
from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
import io

logger = logging.getLogger(__name__)

class ExcelReporter:
    """Generate comprehensive Excel reports for AWS cost optimization"""
    
    # Color scheme
    COLORS = {
        'header': 'FF232F3E',      # AWS Dark Blue
        'subheader': 'FFFF9900',   # AWS Orange
        'savings': 'FF2ECC71',     # Green
        'warning': 'FFF39C12',     # Orange
        'danger': 'FFE74C3C',      # Red
        'info': 'FF3498DB',        # Blue
        'light_gray': 'FFF5F5F5',  # Light Gray
        'medium_gray': 'FFE0E0E0'  # Medium Gray
    }
    
    def __init__(self):
        """Initialize Excel Reporter"""
        self.workbook = None
        self.styles = {}
        
    def generate_comprehensive_report(self,
                                    optimization_result: Dict[str, Any],
                                    output_file: str = 'aws_cost_optimization_report.xlsx'):
        """
        Generate comprehensive Excel report from optimization results
        
        Args:
            optimization_result: Complete optimization results dictionary
            output_file: Output Excel file path
        """
        logger.info(f"Generating comprehensive Excel report: {output_file}")
        
        # Create workbook
        self.workbook = Workbook()
        self._setup_styles()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create sheets
        self._create_executive_summary(optimization_result)
        self._create_savings_dashboard(optimization_result)
        self._create_ec2_analysis(optimization_result.get('details', {}).get('ec2', {}))
        self._create_network_analysis(optimization_result.get('details', {}).get('network', {}))
        self._create_rds_analysis(optimization_result.get('details', {}).get('rds', {}))
        self._create_s3_analysis(optimization_result.get('details', {}).get('s3', {}))
        self._create_ri_analysis(optimization_result.get('details', {}).get('reserved_instances', {}))
        self._create_anomaly_report(optimization_result.get('details', {}).get('anomalies', []))
        self._create_implementation_guide(optimization_result)
        self._create_cost_trends(optimization_result)
        
        # Save workbook
        self.workbook.save(output_file)
        logger.info(f"Report saved to {output_file}")
        
    def _setup_styles(self):
        """Setup reusable styles"""
        # Header style
        self.styles['header'] = {
            'font': Font(size=16, bold=True, color='FFFFFFFF'),
            'fill': PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(bottom=Side(style='medium'))
        }
        
        # Subheader style
        self.styles['subheader'] = {
            'font': Font(size=12, bold=True),
            'fill': PatternFill(start_color=self.COLORS['light_gray'], end_color=self.COLORS['light_gray'], fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(bottom=Side(style='thin'))
        }
        
        # Table header style
        self.styles['table_header'] = {
            'font': Font(bold=True, color='FFFFFFFF'),
            'fill': PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        # Currency style
        self.styles['currency'] = {
            'number_format': '$#,##0.00',
            'alignment': Alignment(horizontal='right')
        }
        
        # Percentage style
        self.styles['percentage'] = {
            'number_format': '0.0%',
            'alignment': Alignment(horizontal='right')
        }
        
        # Savings style
        self.styles['savings'] = {
            'font': Font(bold=True, color=self.COLORS['savings']),
            'number_format': '$#,##0.00',
            'alignment': Alignment(horizontal='right')
        }
        
    def _create_executive_summary(self, optimization_result: Dict[str, Any]):
        """Create executive summary sheet"""
        ws = self.workbook.create_sheet('Executive Summary')
        
        # Title
        ws['A1'] = 'AWS Cost Optimization Report'
        self._apply_style(ws['A1'], self.styles['header'])
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        # Report metadata
        ws['A3'] = 'Report Generated:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Analysis Period:'
        ws['B4'] = '90 days'
        
        # Key metrics
        row = 6
        ws[f'A{row}'] = 'Key Metrics'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'A{row}:F{row}')
        
        metrics = [
            ('Total Monthly Savings Identified', optimization_result.get('total_monthly_savings', 0)),
            ('Total Annual Savings Identified', optimization_result.get('total_annual_savings', 0)),
            ('Number of Recommendations', optimization_result.get('recommendations_count', 0)),
            ('Cost Anomalies Detected', optimization_result.get('anomalies_detected', 0)),
            ('Auto-Remediation Tasks Created', optimization_result.get('auto_remediation_tasks', 0)),
            ('Analysis Execution Time', f"{optimization_result.get('execution_time', 0):.2f} seconds")
        ]
        
        row += 2
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'C{row}'] = value
            if isinstance(value, (int, float)) and 'Savings' in metric:
                self._apply_style(ws[f'C{row}'], self.styles['savings'])
            row += 1
        
        # Savings breakdown
        row += 2
        ws[f'A{row}'] = 'Savings Breakdown by Service'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'A{row}:F{row}')
        
        # Create savings table
        row += 2
        headers = ['Service', 'Monthly Savings', 'Annual Savings', '% of Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['table_header'])
        
        details = optimization_result.get('details', {})
        total_monthly = optimization_result.get('total_monthly_savings', 1)  # Avoid division by zero
        
        savings_data = [
            ('EC2 Instances', details.get('ec2', {}).get('total_monthly_savings', 0)),
            ('Network Resources', details.get('network', {}).get('total_monthly_savings', 0)),
            ('RDS Databases', details.get('rds', {}).get('total_monthly_savings', 0)),
            ('S3 Storage', details.get('s3', {}).get('total_monthly_savings', 0)),
            ('Reserved Instances', details.get('reserved_instances', {}).get('total_monthly_savings', 0))
        ]
        
        row += 1
        for service, monthly_savings in savings_data:
            ws[f'A{row}'] = service
            ws[f'B{row}'] = monthly_savings
            ws[f'C{row}'] = monthly_savings * 12
            ws[f'D{row}'] = monthly_savings / total_monthly if total_monthly > 0 else 0
            
            self._apply_style(ws[f'B{row}'], self.styles['currency'])
            self._apply_style(ws[f'C{row}'], self.styles['currency'])
            self._apply_style(ws[f'D{row}'], self.styles['percentage'])
            row += 1
        
        # Add total row
        ws[f'A{row}'] = 'TOTAL'
        ws[f'B{row}'] = f'=SUM(B{row-len(savings_data)}:B{row-1})'
        ws[f'C{row}'] = f'=SUM(C{row-len(savings_data)}:C{row-1})'
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
        self._apply_style(ws[f'B{row}'], self.styles['savings'])
        self._apply_style(ws[f'C{row}'], self.styles['savings'])
        
        # Add pie chart
        if any(data[1] > 0 for data in savings_data):
            self._add_pie_chart(ws, savings_data, 'F6', 'Savings Distribution by Service')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
    def _create_savings_dashboard(self, optimization_result: Dict[str, Any]):
        """Create visual savings dashboard"""
        ws = self.workbook.create_sheet('Savings Dashboard')
        
        # Title
        ws['A1'] = 'Cost Optimization Dashboard'
        self._apply_style(ws['A1'], self.styles['header'])
        ws.merge_cells('A1:H1')
        
        # Quick wins section
        row = 3
        ws[f'A{row}'] = 'Quick Wins (Low Risk, High Savings)'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'A{row}:H{row}')
        
        # Extract quick wins from all recommendations
        quick_wins = self._extract_quick_wins(optimization_result)
        
        row += 2
        headers = ['Resource', 'Type', 'Action', 'Monthly Savings', 'Risk', 'Confidence']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['table_header'])
        
        row += 1
        for win in quick_wins[:10]:  # Top 10 quick wins
            ws[f'A{row}'] = win.get('resource_id', 'N/A')
            ws[f'B{row}'] = win.get('resource_type', 'N/A')
            ws[f'C{row}'] = win.get('action', 'N/A')
            ws[f'D{row}'] = win.get('monthly_savings', 0)
            ws[f'E{row}'] = win.get('risk_level', 'N/A')
            ws[f'F{row}'] = win.get('confidence', 0)
            
            self._apply_style(ws[f'D{row}'], self.styles['currency'])
            self._apply_style(ws[f'F{row}'], self.styles['percentage'])
            
            # Color code risk
            if win.get('risk_level') == 'low':
                ws[f'E{row}'].font = Font(color=self.COLORS['savings'])
            elif win.get('risk_level') == 'high':
                ws[f'E{row}'].font = Font(color=self.COLORS['danger'])
            
            row += 1
        
        # Implementation timeline
        row += 3
        ws[f'A{row}'] = 'Recommended Implementation Timeline'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'A{row}:H{row}')
        
        timeline = [
            ('Week 1', 'Implement all "stop idle instance" recommendations', '$XX,XXX'),
            ('Week 2', 'Execute rightsizing for development environments', '$XX,XXX'),
            ('Week 3', 'Enable S3 lifecycle policies and Intelligent-Tiering', '$XX,XXX'),
            ('Week 4', 'Review and purchase Reserved Instances', '$XX,XXX'),
            ('Month 2', 'Implement network optimizations', '$XX,XXX'),
            ('Month 3', 'Complete database optimizations', '$XX,XXX')
        ]
        
        row += 2
        headers = ['Timeline', 'Actions', 'Expected Savings']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['table_header'])
        
        row += 1
        for phase, actions, savings in timeline:
            ws[f'A{row}'] = phase
            ws[f'B{row}'] = actions
            ws[f'C{row}'] = savings
            row += 1
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['B'].width = 30
        
    def _create_ec2_analysis(self, ec2_data: Dict[str, Any]):
        """Create EC2 optimization analysis sheet"""
        if not ec2_data or not ec2_data.get('recommendations'):
            return
            
        ws = self.workbook.create_sheet('EC2 Optimization')
        
        # Title
        ws['A1'] = 'EC2 Instance Optimization Analysis'
        self._apply_style(ws['A1'], self.styles['header'])
        ws.merge_cells('A1:I1')
        
        # Summary
        row = 3
        ws[f'A{row}'] = f"Total Monthly Savings: ${ec2_data.get('total_monthly_savings', 0):,.2f}"
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.COLORS['savings'])
        
        # Recommendations table
        row = 5
        ws[f'A{row}'] = 'EC2 Optimization Recommendations'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'A{row}:I{row}')
        
        row += 2
        headers = ['Instance ID', 'Type', 'Region', 'Action', 'Current Cost', 'New Cost', 
                  'Monthly Savings', 'Risk', 'Reason']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['table_header'])
        
        row += 1
        for rec in ec2_data.get('recommendations', []):
            ws[f'A{row}'] = rec.get('instance_id', 'N/A')
            ws[f'B{row}'] = rec.get('instance_type', 'N/A')
            ws[f'C{row}'] = rec.get('region', 'N/A')
            ws[f'D{row}'] = rec.get('action', 'N/A')
            ws[f'E{row}'] = rec.get('current_monthly_cost', 0)
            ws[f'F{row}'] = rec.get('recommended_monthly_cost', 0)
            ws[f'G{row}'] = rec.get('monthly_savings', 0)
            ws[f'H{row}'] = rec.get('risk_level', 'N/A')
            ws[f'I{row}'] = rec.get('reason', 'N/A')
            
            # Apply styles
            for col in ['E', 'F', 'G']:
                self._apply_style(ws[f'{col}{row}'], self.styles['currency'])
            
            # Highlight savings
            ws[f'G{row}'].font = Font(bold=True, color=self.COLORS['savings'])
            
            # Color code actions
            action_colors = {
                'stop': self.COLORS['danger'],
                'rightsize': self.COLORS['warning'],
                'schedule': self.COLORS['info']
            }
            if rec.get('action') in action_colors:
                ws[f'D{row}'].font = Font(color=action_colors[rec.get('action')])
            
            row += 1
        
        # Add summary by action type
        row += 2
        ws[f'A{row}'] = 'Summary by Action Type'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        
        # Group recommendations by action
        action_summary = {}
        for rec in ec2_data.get('recommendations', []):
            action = rec.get('action', 'unknown')
            if action not in action_summary:
                action_summary[action] = {'count': 0, 'savings': 0}
            action_summary[action]['count'] += 1
            action_summary[action]['savings'] += rec.get('monthly_savings', 0)
        
        row += 2
        ws[f'A{row}'] = 'Action'
        ws[f'B{row}'] = 'Count'
        ws[f'C{row}'] = 'Monthly Savings'
        for col in ['A', 'B', 'C']:
            self._apply_style(ws.cell(row=row, column=ord(col)-64), self.styles['table_header'])
        
        row += 1
        for action, data in action_summary.items():
            ws[f'A{row}'] = action.title()
            ws[f'B{row}'] = data['count']
            ws[f'C{row}'] = data['savings']
            self._apply_style(ws[f'C{row}'], self.styles['currency'])
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['I'].width = 50
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
    
    def _create_network_analysis(self, network_data: Dict[str, Any]):
        """Create network optimization analysis sheet"""
        if not network_data or not network_data.get('recommendations'):
            return
            
        ws = self.workbook.create_sheet('Network Optimization')
        
        # Title
        ws['A1'] = 'Network Cost Optimization Analysis'
        self._apply_style(ws['A1'], self.styles['header'])
        ws.merge_cells('A1:H1')
        
        # Add content similar to EC2 sheet
        # ... (implement similar structure)
        
    def _create_rds_analysis(self, rds_data: Dict[str, Any]):
        """Create RDS optimization analysis sheet"""
        # Similar implementation to EC2
        pass
        
    def _create_s3_analysis(self, s3_data: Dict[str, Any]):
        """Create S3 optimization analysis sheet"""
        # Similar implementation to EC2
        pass
        
    def _create_ri_analysis(self, ri_data: Dict[str, Any]):
        """Create Reserved Instance analysis sheet"""
        if not ri_data:
            return
            
        ws = self.workbook.create_sheet('Reserved Instances')
        
        # Title
        ws['A1'] = 'Reserved Instance & Savings Plan Analysis'
        self._apply_style(ws['A1'], self.styles['header'])
        ws.merge_cells('A1:J1')
        
        # Current utilization
        row = 3
        ws[f'A{row}'] = 'Current Utilization'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        
        row += 2
        utilization = ri_data.get('current_utilization', {})
        ws[f'A{row}'] = 'Reserved Instance Utilization:'
        ws[f'C{row}'] = f"{utilization.get('reserved_instances', {}).get('total_utilization', 0):.1f}%"
        row += 1
        ws[f'A{row}'] = 'Savings Plan Utilization:'
        ws[f'C{row}'] = f"{utilization.get('savings_plans', {}).get('total_utilization', 0):.1f}%"
        
        # RI Recommendations
        row += 3
        ws[f'A{row}'] = 'Reserved Instance Recommendations'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'A{row}:J{row}')
        
        if ri_data.get('ri_recommendations'):
            row += 2
            headers = ['Instance Type', 'Region', 'Quantity', 'Term', 'Payment', 'Class',
                      'Monthly Savings', 'Annual Savings', 'ROI (months)', 'Confidence']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                self._apply_style(cell, self.styles['table_header'])
            
            row += 1
            for rec in ri_data.get('ri_recommendations', [])[:20]:  # Top 20
                ws[f'A{row}'] = rec.get('instance_type', 'N/A')
                ws[f'B{row}'] = rec.get('region', 'N/A')
                ws[f'C{row}'] = rec.get('quantity', 0)
                ws[f'D{row}'] = f"{rec.get('term_length', 0)} Year"
                ws[f'E{row}'] = rec.get('payment_option', 'N/A').replace('_', ' ').title()
                ws[f'F{row}'] = rec.get('offering_class', 'N/A').title()
                ws[f'G{row}'] = rec.get('monthly_savings', 0)
                ws[f'H{row}'] = rec.get('annual_savings', 0)
                ws[f'I{row}'] = rec.get('roi_months', 0)
                ws[f'J{row}'] = rec.get('confidence_score', 0)
                
                # Apply styles
                self._apply_style(ws[f'G{row}'], self.styles['currency'])
                self._apply_style(ws[f'H{row}'], self.styles['currency'])
                self._apply_style(ws[f'J{row}'], self.styles['percentage'])
                
                row += 1
        
        # Adjust column widths
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 15
            
    def _create_anomaly_report(self, anomalies: List[Dict[str, Any]]):
        """Create cost anomaly report sheet"""
        if not anomalies:
            return
            
        ws = self.workbook.create_sheet('Cost Anomalies')
        
        # Title
        ws['A1'] = 'Cost Anomaly Detection Report'
        self._apply_style(ws['A1'], self.styles['header'])
        ws.merge_cells('A1:J1')
        
        # Anomaly table
        row = 3
        headers = ['Service', 'Type', 'Severity', 'Current Cost', 'Expected Cost',
                  'Impact', 'Increase %', 'Detection Date', 'Confidence', 'Probable Causes']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['table_header'])
        
        row += 1
        for anomaly in anomalies:
            ws[f'A{row}'] = anomaly.get('service', 'N/A')
            ws[f'B{row}'] = anomaly.get('anomaly_type', 'N/A').replace('_', ' ').title()
            ws[f'C{row}'] = anomaly.get('severity', 'N/A').upper()
            ws[f'D{row}'] = anomaly.get('current_daily_cost', 0)
            ws[f'E{row}'] = anomaly.get('expected_daily_cost', 0)
            ws[f'F{row}'] = anomaly.get('cost_impact', 0)
            ws[f'G{row}'] = anomaly.get('percentage_increase', 0) / 100
            ws[f'H{row}'] = anomaly.get('detection_date', 'N/A')
            ws[f'I{row}'] = anomaly.get('confidence_score', 0)
            ws[f'J{row}'] = '; '.join(anomaly.get('probable_causes', []))[:100]
            
            # Apply styles
            for col in ['D', 'E', 'F']:
                self._apply_style(ws[f'{col}{row}'], self.styles['currency'])
            self._apply_style(ws[f'G{row}'], self.styles['percentage'])
            self._apply_style(ws[f'I{row}'], self.styles['percentage'])
            
            # Color code severity
            severity_colors = {
                'CRITICAL': self.COLORS['danger'],
                'HIGH': self.COLORS['warning'],
                'MEDIUM': self.COLORS['info']
            }
            if anomaly.get('severity', '').upper() in severity_colors:
                ws[f'C{row}'].font = Font(bold=True, color=severity_colors[anomaly.get('severity', '').upper()])
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['J'].width = 50
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 15
            
    def _create_implementation_guide(self, optimization_result: Dict[str, Any]):
        """Create implementation guide sheet"""
        ws = self.workbook.create_sheet('Implementation Guide')
        
        # Title
        ws['A1'] = 'Implementation Guide'
        self._apply_style(ws['A1'], self.styles['header'])
        ws.merge_cells('A1:D1')
        
        # Phase 1: Quick Wins
        row = 3
        ws[f'A{row}'] = 'Phase 1: Quick Wins (Week 1)'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        quick_wins = [
            ('Stop idle EC2 instances', 'Low', 'Run provided CLI commands or use console'),
            ('Delete unattached EBS volumes', 'Low', 'Review list and delete after snapshot'),
            ('Release unused Elastic IPs', 'Low', 'Verify not needed and release'),
            ('Enable S3 Intelligent-Tiering', 'Low', 'Apply lifecycle policies provided')
        ]
        
        headers = ['Action', 'Risk', 'Implementation Steps']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['table_header'])
        
        row += 1
        for action, risk, steps in quick_wins:
            ws[f'A{row}'] = action
            ws[f'B{row}'] = risk
            ws[f'C{row}'] = steps
            row += 1
        
        # Phase 2: Rightsizing
        row += 2
        ws[f'A{row}'] = 'Phase 2: Rightsizing (Week 2-3)'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        rightsizing = [
            ('EC2 instance rightsizing', 'Medium', 'Test in dev first, schedule maintenance window'),
            ('RDS instance rightsizing', 'Medium', 'Create snapshot, modify during maintenance'),
            ('Implement instance scheduling', 'Low', 'Tag resources and enable scheduler')
        ]
        
        for action, risk, steps in rightsizing:
            ws[f'A{row}'] = action
            ws[f'B{row}'] = risk
            ws[f'C{row}'] = steps
            row += 1
        
        # Best practices
        row += 3
        ws[f'A{row}'] = 'Best Practices'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        practices = [
            '1. Always create backups before making changes',
            '2. Test changes in non-production environments first',
            '3. Implement changes during maintenance windows',
            '4. Monitor performance after changes',
            '5. Have a rollback plan ready',
            '6. Document all changes made',
            '7. Communicate with stakeholders'
        ]
        
        for practice in practices:
            ws[f'A{row}'] = practice
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        
    def _create_cost_trends(self, optimization_result: Dict[str, Any]):
        """Create cost trends and projections sheet"""
        ws = self.workbook.create_sheet('Cost Trends')
        
        # Title
        ws['A1'] = 'Cost Trends & Projections'
        self._apply_style(ws['A1'], self.styles['header'])
        ws.merge_cells('A1:F1')
        
        # Current vs Optimized costs
        row = 3
        ws[f'A{row}'] = 'Cost Projections'
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        
        # Create projection table
        row += 2
        headers = ['Month', 'Current Cost', 'Optimized Cost', 'Savings', 'Cumulative Savings']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['table_header'])
        
        # Generate 12-month projection
        current_monthly = 100000  # Example base cost
        monthly_savings = optimization_result.get('total_monthly_savings', 0)
        
        row += 1
        cumulative = 0
        for month in range(1, 13):
            ws[f'A{row}'] = f'Month {month}'
            ws[f'B{row}'] = current_monthly
            ws[f'C{row}'] = current_monthly - monthly_savings
            ws[f'D{row}'] = monthly_savings
            cumulative += monthly_savings
            ws[f'E{row}'] = cumulative
            
            for col in ['B', 'C', 'D', 'E']:
                self._apply_style(ws[f'{col}{row}'], self.styles['currency'])
            
            row += 1
        
        # Add line chart for visual representation
        if row > 7:  # Only add chart if we have data
            chart = LineChart()
            chart.title = "Cost Trend - Current vs Optimized"
            chart.style = 10
            chart.y_axis.title = 'Monthly Cost ($)'
            chart.x_axis.title = 'Month'
            
            # Data for current cost
            data1 = Reference(ws, min_col=2, min_row=5, max_col=2, max_row=row-1)
            chart.add_data(data1, titles_from_data=False)
            chart.series[0].title = "Current Cost"
            
            # Data for optimized cost
            data2 = Reference(ws, min_col=3, min_row=5, max_col=3, max_row=row-1)
            chart.add_data(data2, titles_from_data=False)
            chart.series[1].title = "Optimized Cost"
            
            # Categories
            cats = Reference(ws, min_col=1, min_row=6, max_row=row-1)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "G3")
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
            
    def _apply_style(self, cell, style_dict):
        """Apply style dictionary to a cell"""
        for attr, value in style_dict.items():
            setattr(cell, attr, value)
            
    def _add_pie_chart(self, ws, data, position, title):
        """Add a pie chart to worksheet"""
        pie = PieChart()
        pie.title = title
        
        # Prepare data for chart
        row_start = ws.max_row + 2
        col_labels = 1
        col_values = 2
        
        for i, (label, value) in enumerate(data):
            ws.cell(row=row_start + i, column=col_labels, value=label)
            ws.cell(row=row_start + i, column=col_values, value=value)
        
        # Add data to chart
        labels = Reference(ws, min_col=col_labels, min_row=row_start, max_row=row_start + len(data) - 1)
        data_ref = Reference(ws, min_col=col_values, min_row=row_start, max_row=row_start + len(data) - 1)
        
        pie.add_data(data_ref)
        pie.set_categories(labels)
        
        # Style the chart
        pie.height = 10
        pie.width = 15
        
        ws.add_chart(pie, position)
        
        # Hide the data used for chart
        for i in range(len(data)):
            ws.row_dimensions[row_start + i].hidden = True
            
    def _extract_quick_wins(self, optimization_result: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract quick win recommendations from all services"""
        quick_wins = []
        
        # Define what makes a quick win
        def is_quick_win(rec):
            return (
                rec.get('risk_level', '').lower() == 'low' and
                rec.get('confidence', 0) >= 0.8 and
                rec.get('monthly_savings', 0) > 50
            )
        
        # Extract from each service
        details = optimization_result.get('details', {})
        
        # EC2 quick wins
        for rec in details.get('ec2', {}).get('recommendations', []):
            if is_quick_win(rec):
                quick_wins.append({
                    'resource_id': rec.get('instance_id'),
                    'resource_type': 'EC2',
                    'action': rec.get('action'),
                    'monthly_savings': rec.get('monthly_savings'),
                    'risk_level': rec.get('risk_level'),
                    'confidence': rec.get('confidence')
                })
        
        # Network quick wins
        for rec in details.get('network', {}).get('recommendations', []):
            if is_quick_win(rec):
                quick_wins.append({
                    'resource_id': rec.get('resource_id'),
                    'resource_type': rec.get('resource_type'),
                    'action': rec.get('recommended_action'),
                    'monthly_savings': rec.get('estimated_monthly_savings'),
                    'risk_level': rec.get('risk_level'),
                    'confidence': rec.get('confidence')
                })
        
        # Sort by savings
        quick_wins.sort(key=lambda x: x.get('monthly_savings', 0), reverse=True)
        
        return quick_wins
    
    def generate_simple_report(self, 
                             recommendations: List[Dict[str, Any]],
                             service_name: str,
                             output_file: str):
        """Generate a simple report for a single service"""
        workbook = Workbook()
        ws = workbook.active
        ws.title = f'{service_name} Recommendations'
        
        # Title
        ws['A1'] = f'{service_name} Optimization Recommendations'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Add recommendations
        row = 3
        headers = ['Resource ID', 'Action', 'Monthly Savings', 'Risk Level', 'Reason']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFCCCCCC', end_color='FFCCCCCC', fill_type='solid')
        
        row += 1
        total_savings = 0
        
        for rec in recommendations:
            ws[f'A{row}'] = rec.get('resource_id', 'N/A')
            ws[f'B{row}'] = rec.get('action', 'N/A')
            ws[f'C{row}'] = rec.get('monthly_savings', 0)
            ws[f'D{row}'] = rec.get('risk_level', 'N/A')
            ws[f'E{row}'] = rec.get('reason', 'N/A')
            
            # Format currency
            ws[f'C{row}'].number_format = '$#,##0.00'
            
            total_savings += rec.get('monthly_savings', 0)
            row += 1
        
        # Add total
        row += 1
        ws[f'B{row}'] = 'TOTAL'
        ws[f'C{row}'] = total_savings
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].number_format = '$#,##0.00'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50
        
        workbook.save(output_file)
        logger.info(f"Simple report saved to {output_file}")