"""
Periodic Resource Detection Module

Identifies resources with periodic usage patterns (monthly, quarterly, yearly)
to prevent optimization of critical batch jobs and seasonal workloads.
"""

import numpy as np
from scipy import signal, stats
from typing import Dict, List, Tuple, Optional, Any
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from enum import Enum
import pandas as pd
import boto3
from collections import defaultdict
import logging

logger = logging.getLogger(__name__)


class PeriodType(Enum):
    """Types of periodic patterns"""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"
    CUSTOM = "custom"
    NONE = "none"


@dataclass
class PeriodicPattern:
    """Represents a detected periodic pattern"""
    period_type: PeriodType
    period_days: float
    confidence: float
    peak_times: List[datetime]
    description: str
    next_expected_usage: Optional[datetime] = None
    is_critical_period: bool = False
    business_context: Optional[str] = None
    
    
@dataclass
class ResourcePeriodicity:
    """Periodicity analysis results for a resource"""
    resource_id: str
    resource_type: str
    patterns: List[PeriodicPattern]
    usage_classification: str  # 'continuous', 'periodic', 'sporadic', 'idle'
    recommendation: str
    risk_score: float  # 0-1, higher means riskier to optimize
    analysis_period_days: int
    last_analyzed: datetime


class PeriodicResourceDetector:
    """Detect resources with periodic usage patterns"""
    
    # Known business periods
    BUSINESS_PERIODS = {
        'month_end': (28, 31),
        'quarter_end': (89, 92),
        'year_end': (364, 366),
        'bi_weekly': (13, 15),
        'fiscal_quarter': (91, 91)  # Exactly 13 weeks
    }
    
    # Minimum data points for reliable analysis
    MIN_DATA_POINTS = {
        PeriodType.DAILY: 14,
        PeriodType.WEEKLY: 8,
        PeriodType.MONTHLY: 3,
        PeriodType.QUARTERLY: 4,
        PeriodType.YEARLY: 2
    }
    
    def __init__(self, 
                 lookback_days: int = 365,
                 confidence_threshold: float = 0.7,
                 cloudwatch_client=None):
        """
        Initialize periodic detector
        
        Args:
            lookback_days: Days of history to analyze
            confidence_threshold: Minimum confidence for pattern detection
            cloudwatch_client: Boto3 CloudWatch client
        """
        self.lookback_days = lookback_days
        self.confidence_threshold = confidence_threshold
        self.cloudwatch = cloudwatch_client or boto3.client('cloudwatch')
        
    def analyze_resource(self, 
                        resource_id: str,
                        resource_type: str,
                        metric_name: str = 'CPUUtilization',
                        namespace: str = 'AWS/EC2') -> ResourcePeriodicity:
        """
        Analyze a single resource for periodic patterns
        
        Args:
            resource_id: Resource identifier
            resource_type: Type of resource (ec2, rds, etc)
            metric_name: CloudWatch metric to analyze
            namespace: CloudWatch namespace
            
        Returns:
            ResourcePeriodicity object with analysis results
        """
        logger.info(f"Analyzing periodicity for {resource_type} {resource_id}")
        
        # Fetch metric data
        time_series = self._fetch_metric_data(
            resource_id, resource_type, metric_name, namespace
        )
        
        if len(time_series) < 30:
            logger.warning(f"Insufficient data for {resource_id}, only {len(time_series)} points")
            return self._create_insufficient_data_result(resource_id, resource_type)
        
        # Detect patterns using multiple methods
        patterns = []
        
        # 1. Fourier analysis for frequency detection
        fourier_patterns = self._fourier_analysis(time_series)
        patterns.extend(fourier_patterns)
        
        # 2. Statistical seasonality tests
        seasonal_patterns = self._seasonal_decomposition(time_series)
        patterns.extend(seasonal_patterns)
        
        # 3. Peak detection and pattern matching
        peak_patterns = self._peak_pattern_analysis(time_series)
        patterns.extend(peak_patterns)
        
        # 4. Business calendar correlation
        business_patterns = self._business_calendar_correlation(time_series)
        patterns.extend(business_patterns)
        
        # Deduplicate and rank patterns
        patterns = self._consolidate_patterns(patterns)
        
        # Classify usage and generate recommendation
        classification = self._classify_usage(patterns, time_series)
        recommendation = self._generate_recommendation(classification, patterns)
        risk_score = self._calculate_risk_score(patterns, classification)
        
        return ResourcePeriodicity(
            resource_id=resource_id,
            resource_type=resource_type,
            patterns=patterns,
            usage_classification=classification,
            recommendation=recommendation,
            risk_score=risk_score,
            analysis_period_days=self.lookback_days,
            last_analyzed=datetime.utcnow()
        )
    
    def _fetch_metric_data(self, 
                          resource_id: str,
                          resource_type: str,
                          metric_name: str,
                          namespace: str) -> pd.Series:
        """Fetch CloudWatch metrics for analysis"""
        end_time = datetime.utcnow()
        start_time = end_time - timedelta(days=self.lookback_days)
        
        # Determine dimension name based on resource type
        dimension_map = {
            'ec2': 'InstanceId',
            'rds': 'DBInstanceIdentifier',
            'elasticache': 'CacheClusterId',
            'ecs': 'ServiceName',
            'lambda': 'FunctionName'
        }
        
        dimension_name = dimension_map.get(resource_type.lower(), 'InstanceId')
        
        # Fetch metric statistics
        try:
            response = self.cloudwatch.get_metric_statistics(
                Namespace=namespace,
                MetricName=metric_name,
                Dimensions=[{
                    'Name': dimension_name,
                    'Value': resource_id
                }],
                StartTime=start_time,
                EndTime=end_time,
                Period=3600,  # 1 hour granularity
                Statistics=['Maximum', 'Average']
            )
        except Exception as e:
            logger.error(f"Failed to fetch CloudWatch metrics for {resource_id}: {e}")
            return pd.Series()
        
        # Convert to pandas series for analysis
        if not response['Datapoints']:
            return pd.Series()
            
        df = pd.DataFrame(response['Datapoints'])
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])
        df = df.sort_values('Timestamp')
        df.set_index('Timestamp', inplace=True)
        
        # Use maximum values for peak detection
        return df['Maximum'].fillna(0)
    
    def _fourier_analysis(self, time_series: pd.Series) -> List[PeriodicPattern]:
        """Use FFT to detect periodic patterns"""
        patterns = []
        
        if len(time_series) < 48:  # Need at least 2 days of hourly data
            return patterns
        
        # Detrend the data
        values = time_series.values
        detrended = signal.detrend(values)
        
        # Apply FFT
        fft_result = np.fft.fft(detrended)
        frequencies = np.fft.fftfreq(len(detrended), d=1.0)  # Assuming hourly data
        
        # Get power spectrum
        power = np.abs(fft_result) ** 2
        
        # Find peaks in power spectrum
        peaks, properties = signal.find_peaks(
            power[:len(power)//2],  # Only positive frequencies
            height=np.mean(power) * 2,  # Peaks above 2x mean power
            distance=5  # Minimum distance between peaks
        )
        
        for peak_idx in peaks[:5]:  # Top 5 peaks
            frequency = frequencies[peak_idx]
            if frequency > 0:
                period_hours = 1.0 / frequency
                period_days = period_hours / 24.0
                
                # Map to known periods
                period_type, confidence = self._map_to_period_type(period_days)
                
                if confidence >= self.confidence_threshold:
                    patterns.append(PeriodicPattern(
                        period_type=period_type,
                        period_days=period_days,
                        confidence=confidence,
                        peak_times=self._find_peak_times(time_series, period_hours),
                        description=f"Fourier analysis detected {period_type.value} pattern"
                    ))
        
        return patterns
    
    def _seasonal_decomposition(self, time_series: pd.Series) -> List[PeriodicPattern]:
        """Use statistical decomposition to find seasonal patterns"""
        patterns = []
        
        # Try different seasonal periods
        test_periods = [
            (24, PeriodType.DAILY),      # Daily
            (24 * 7, PeriodType.WEEKLY),  # Weekly
            (24 * 30, PeriodType.MONTHLY) # Monthly
        ]
        
        for period_hours, period_type in test_periods:
            if len(time_series) < period_hours * 2:
                continue
                
            try:
                # Use statsmodels seasonal decompose if available
                from statsmodels.tsa.seasonal import seasonal_decompose
                
                decomposition = seasonal_decompose(
                    time_series,
                    model='additive',
                    period=period_hours
                )
                
                # Calculate strength of seasonality
                seasonal_strength = np.var(decomposition.seasonal) / np.var(time_series)
                
                if seasonal_strength > 0.1:  # Significant seasonal component
                    confidence = min(seasonal_strength * 2, 1.0)
                    
                    patterns.append(PeriodicPattern(
                        period_type=period_type,
                        period_days=period_hours / 24.0,
                        confidence=confidence,
                        peak_times=self._find_peak_times(time_series, period_hours),
                        description=f"Seasonal decomposition found {period_type.value} pattern"
                    ))
                    
            except Exception as e:
                logger.debug(f"Seasonal decomposition failed for period {period_hours}: {e}")
                
        return patterns
    
    def _peak_pattern_analysis(self, time_series: pd.Series) -> List[PeriodicPattern]:
        """Analyze peaks to find recurring patterns"""
        patterns = []
        
        # Find peaks
        values = time_series.values
        mean_val = np.mean(values)
        std_val = np.std(values)
        
        # Peaks are values above mean + 1.5 * std
        peak_threshold = mean_val + 1.5 * std_val
        peaks, _ = signal.find_peaks(values, height=peak_threshold)
        
        if len(peaks) < 3:
            return patterns
        
        # Analyze intervals between peaks
        peak_times = time_series.index[peaks]
        intervals = np.diff(peak_times) / pd.Timedelta(hours=1)  # Convert to hours
        
        # Look for recurring intervals
        for test_interval in [24, 24*7, 24*30, 24*91]:  # Daily, weekly, monthly, quarterly
            matches = np.abs(intervals - test_interval) < test_interval * 0.1  # 10% tolerance
            
            if np.sum(matches) >= 2:  # At least 2 matching intervals
                period_type, confidence = self._map_to_period_type(test_interval / 24)
                confidence *= np.sum(matches) / len(intervals)  # Adjust by match ratio
                
                if confidence >= self.confidence_threshold:
                    patterns.append(PeriodicPattern(
                        period_type=period_type,
                        period_days=test_interval / 24,
                        confidence=confidence,
                        peak_times=list(peak_times),
                        description=f"Peak analysis found recurring {period_type.value} pattern"
                    ))
        
        return patterns
    
    def _business_calendar_correlation(self, time_series: pd.Series) -> List[PeriodicPattern]:
        """Correlate usage with business calendar events"""
        patterns = []
        
        # Group by different time periods
        daily_avg = time_series.groupby(time_series.index.day).mean()
        weekly_avg = time_series.groupby(time_series.index.dayofweek).mean()
        
        # Check for month-end pattern
        month_end_usage = daily_avg[daily_avg.index >= 25].mean()
        other_days_usage = daily_avg[daily_avg.index < 25].mean()
        
        if month_end_usage > other_days_usage * 1.5:
            patterns.append(PeriodicPattern(
                period_type=PeriodType.MONTHLY,
                period_days=30,
                confidence=0.8,
                peak_times=self._find_month_end_peaks(time_series),
                description="Month-end processing pattern detected",
                business_context="Month-end batch processing"
            ))
        
        # Check for weekend pattern
        weekend_usage = weekly_avg[[5, 6]].mean()  # Saturday, Sunday
        weekday_usage = weekly_avg[[0, 1, 2, 3, 4]].mean()
        
        if weekend_usage < weekday_usage * 0.3:
            patterns.append(PeriodicPattern(
                period_type=PeriodType.WEEKLY,
                period_days=7,
                confidence=0.85,
                peak_times=[],
                description="Business hours pattern detected",
                business_context="Weekday business operations"
            ))
        
        # Check for quarter-end pattern
        quarter_end_months = [3, 6, 9, 12]
        quarter_end_data = time_series[time_series.index.month.isin(quarter_end_months)]
        other_months_data = time_series[~time_series.index.month.isin(quarter_end_months)]
        
        if len(quarter_end_data) > 0 and len(other_months_data) > 0:
            if quarter_end_data.mean() > other_months_data.mean() * 1.3:
                patterns.append(PeriodicPattern(
                    period_type=PeriodType.QUARTERLY,
                    period_days=91,
                    confidence=0.75,
                    peak_times=self._find_quarter_end_peaks(time_series),
                    description="Quarter-end pattern detected",
                    business_context="Quarterly reporting or processing"
                ))
        
        return patterns
    
    def _map_to_period_type(self, period_days: float) -> Tuple[PeriodType, float]:
        """Map a period in days to a PeriodType with confidence"""
        period_map = [
            (1, PeriodType.DAILY, 0.02),
            (7, PeriodType.WEEKLY, 0.1),
            (30, PeriodType.MONTHLY, 0.5),
            (91, PeriodType.QUARTERLY, 1.0),
            (365, PeriodType.YEARLY, 2.0)
        ]
        
        for expected_days, period_type, tolerance_ratio in period_map:
            tolerance = expected_days * tolerance_ratio
            if abs(period_days - expected_days) <= tolerance:
                confidence = 1.0 - (abs(period_days - expected_days) / tolerance)
                return period_type, confidence
        
        return PeriodType.CUSTOM, 0.5
    
    def _find_peak_times(self, time_series: pd.Series, period_hours: float) -> List[datetime]:
        """Find times when peaks occur within the period"""
        values = time_series.values
        peaks, _ = signal.find_peaks(values, height=np.mean(values) + np.std(values))
        
        peak_times = time_series.index[peaks].tolist()
        return peak_times[-10:]  # Return last 10 peaks
    
    def _find_month_end_peaks(self, time_series: pd.Series) -> List[datetime]:
        """Find month-end peak times"""
        month_end_data = time_series[time_series.index.day >= 25]
        threshold = month_end_data.mean() + month_end_data.std()
        peaks = month_end_data[month_end_data > threshold]
        return peaks.index.tolist()[-10:]
    
    def _find_quarter_end_peaks(self, time_series: pd.Series) -> List[datetime]:
        """Find quarter-end peak times"""
        quarter_months = [3, 6, 9, 12]
        quarter_data = time_series[
            time_series.index.month.isin(quarter_months) & 
            (time_series.index.day >= 20)
        ]
        threshold = quarter_data.mean() + quarter_data.std()
        peaks = quarter_data[quarter_data > threshold]
        return peaks.index.tolist()[-10:]
    
    def _consolidate_patterns(self, patterns: List[PeriodicPattern]) -> List[PeriodicPattern]:
        """Consolidate similar patterns and remove duplicates"""
        if not patterns:
            return []
        
        # Group by period type
        grouped = defaultdict(list)
        for pattern in patterns:
            grouped[pattern.period_type].append(pattern)
        
        # Keep highest confidence pattern for each type
        consolidated = []
        for period_type, group in grouped.items():
            best_pattern = max(group, key=lambda p: p.confidence)
            
            # Merge information from similar patterns
            best_pattern.confidence = np.mean([p.confidence for p in group])
            best_pattern.description = "; ".join(set(p.description for p in group))
            
            consolidated.append(best_pattern)
        
        return sorted(consolidated, key=lambda p: p.confidence, reverse=True)
    
    def _classify_usage(self, patterns: List[PeriodicPattern], time_series: pd.Series) -> str:
        """Classify overall usage pattern"""
        if not patterns:
            # Check if resource is idle
            if time_series.mean() < 5 and time_series.max() < 10:
                return 'idle'
            return 'sporadic'
        
        # If high confidence periodic patterns exist
        high_confidence_patterns = [p for p in patterns if p.confidence >= 0.8]
        if high_confidence_patterns:
            return 'periodic'
        
        # Check coefficient of variation
        cv = time_series.std() / time_series.mean() if time_series.mean() > 0 else 0
        
        if cv < 0.3:
            return 'continuous'
        elif cv > 1.0:
            return 'sporadic'
        else:
            return 'periodic' if patterns else 'sporadic'
    
    def _generate_recommendation(self, classification: str, patterns: List[PeriodicPattern]) -> str:
        """Generate optimization recommendation based on patterns"""
        recommendations = {
            'continuous': "Safe to optimize. Consider Reserved Instances or Savings Plans.",
            'periodic': "Caution: Periodic usage detected. Consider scheduling or keep active.",
            'sporadic': "Irregular usage. Consider on-demand or event-driven architecture.",
            'idle': "Resource appears idle. Strong candidate for termination."
        }
        
        base_rec = recommendations.get(classification, "Requires manual review.")
        
        # Add pattern-specific recommendations
        if patterns:
            pattern_recs = []
            for pattern in patterns[:2]:  # Top 2 patterns
                if pattern.period_type == PeriodType.MONTHLY:
                    pattern_recs.append("Month-end processing detected - DO NOT terminate")
                elif pattern.period_type == PeriodType.QUARTERLY:
                    pattern_recs.append("Quarterly processing detected - Critical for business")
                elif pattern.period_type == PeriodType.WEEKLY:
                    pattern_recs.append("Weekly pattern - Consider weekend scheduling")
                elif pattern.period_type == PeriodType.DAILY:
                    pattern_recs.append("Daily pattern - Consider off-hours scheduling")
            
            if pattern_recs:
                base_rec += " " + ". ".join(pattern_recs) + "."
        
        return base_rec
    
    def _calculate_risk_score(self, patterns: List[PeriodicPattern], classification: str) -> float:
        """Calculate risk score for optimizing this resource (0-1)"""
        base_scores = {
            'continuous': 0.2,
            'periodic': 0.7,
            'sporadic': 0.5,
            'idle': 0.1
        }
        
        risk_score = base_scores.get(classification, 0.5)
        
        # Adjust based on pattern types
        for pattern in patterns:
            if pattern.period_type in [PeriodType.MONTHLY, PeriodType.QUARTERLY, PeriodType.YEARLY]:
                risk_score = max(risk_score, 0.8)
            if pattern.business_context and 'critical' in pattern.business_context.lower():
                risk_score = max(risk_score, 0.9)
        
        return min(risk_score, 1.0)
    
    def _create_insufficient_data_result(self, resource_id: str, resource_type: str) -> ResourcePeriodicity:
        """Create result for resources with insufficient data"""
        return ResourcePeriodicity(
            resource_id=resource_id,
            resource_type=resource_type,
            patterns=[],
            usage_classification='unknown',
            recommendation='Insufficient data for analysis. Gather more history.',
            risk_score=0.5,
            analysis_period_days=self.lookback_days,
            last_analyzed=datetime.utcnow()
        )
    
    def batch_analyze_resources(self, 
                               resources: List[Dict[str, str]],
                               metric_configs: Optional[Dict[str, Dict[str, str]]] = None) -> Dict[str, ResourcePeriodicity]:
        """
        Analyze multiple resources in batch
        
        Args:
            resources: List of dicts with 'resource_id' and 'resource_type'
            metric_configs: Optional metric configuration per resource type
            
        Returns:
            Dict mapping resource_id to ResourcePeriodicity
        """
        results = {}
        
        # Default metric configurations
        default_configs = {
            'ec2': {'metric_name': 'CPUUtilization', 'namespace': 'AWS/EC2'},
            'rds': {'metric_name': 'CPUUtilization', 'namespace': 'AWS/RDS'},
            'elasticache': {'metric_name': 'CPUUtilization', 'namespace': 'AWS/ElastiCache'},
            'ecs': {'metric_name': 'CPUUtilization', 'namespace': 'AWS/ECS'},
            'lambda': {'metric_name': 'Invocations', 'namespace': 'AWS/Lambda'}
        }
        
        metric_configs = metric_configs or default_configs
        
        for resource in resources:
            resource_id = resource['resource_id']
            resource_type = resource['resource_type']
            
            # Get metric configuration
            config = metric_configs.get(resource_type, default_configs.get('ec2'))
            
            try:
                result = self.analyze_resource(
                    resource_id=resource_id,
                    resource_type=resource_type,
                    metric_name=config['metric_name'],
                    namespace=config['namespace']
                )
                results[resource_id] = result
                
            except Exception as e:
                logger.error(f"Failed to analyze {resource_id}: {e}")
                results[resource_id] = self._create_insufficient_data_result(
                    resource_id, resource_type
                )
        
        return results
    
    def export_analysis_report(self, 
                              results: Dict[str, ResourcePeriodicity],
                              output_file: str = 'periodic_analysis_report.xlsx'):
        """Export analysis results to Excel report"""
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Prepare data for export
        report_data = []
        for resource_id, analysis in results.items():
            row = {
                'Resource ID': resource_id,
                'Resource Type': analysis.resource_type,
                'Usage Classification': analysis.usage_classification,
                'Risk Score': f"{analysis.risk_score:.2f}",
                'Primary Pattern': analysis.patterns[0].period_type.value if analysis.patterns else 'None',
                'Confidence': f"{analysis.patterns[0].confidence:.2f}" if analysis.patterns else 'N/A',
                'Recommendation': analysis.recommendation,
                'Analysis Date': analysis.last_analyzed.strftime('%Y-%m-%d %H:%M')
            }
            report_data.append(row)
        
        # Create DataFrame and Excel writer
        df = pd.DataFrame(report_data)
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Periodic Analysis', index=False)
            
            # Format the worksheet
            worksheet = writer.sheets['Periodic Analysis']
            
            # Header formatting
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Conditional formatting for risk scores
            for row in range(2, len(df) + 2):
                risk_cell = worksheet[f'D{row}']
                if risk_cell.value is not None and str(risk_cell.value).replace('.', '').replace('-', '').isdigit():
                    risk_value = float(risk_cell.value)
                else:
                    continue
                
                if risk_value >= 0.8:
                    risk_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif risk_value >= 0.6:
                    risk_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    risk_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                for cell in column:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        
        logger.info(f"Periodic analysis report exported to {output_file}")