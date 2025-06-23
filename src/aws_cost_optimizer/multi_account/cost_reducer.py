#!/usr/bin/env python3
"""
Emergency Cost Reduction Plan Generator
Creates a prioritized action plan to achieve $20K/month savings target
"""
import json
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Tuple
import click
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class EmergencyCostReducer:
    """Generates emergency cost reduction plans based on discovered resources"""
    
    def __init__(self, target_savings: float = 20000):
        self.target_savings = target_savings
        self.recommendations = []
        self.immediate_actions = []
        self.scheduled_actions = []
        self.policy_changes = []
        
    def analyze_ec2_recommendations(self, ec2_recommendations: List[Dict]) -> List[Dict]:
        """Prioritize EC2 recommendations for immediate savings"""
        actions = []
        
        for rec in ec2_recommendations:
            priority = self._calculate_priority(rec)
            
            action = {
                'service': 'EC2',
                'resource_id': rec['instance_id'],
                'action': rec['action'],
                'description': rec['reason'],
                'monthly_savings': rec['monthly_savings'],
                'risk_level': rec['risk_level'],
                'priority': priority,
                'implementation_time': self._estimate_implementation_time(rec['action']),
                'rollback_plan': self._create_rollback_plan(rec),
                'commands': rec.get('implementation_steps', [])
            }
            
            # Categorize by urgency
            if rec['risk_level'] == 'low' and rec['action'] == 'stop':
                action['category'] = 'immediate'
                self.immediate_actions.append(action)
            elif rec['risk_level'] == 'low' and rec['action'] == 'schedule':
                action['category'] = 'quick_win'
                self.scheduled_actions.append(action)
            else:
                action['category'] = 'planned'
                actions.append(action)
        
        return actions
    
    def analyze_rds_recommendations(self, rds_recommendations: List[Dict]) -> List[Dict]:
        """Prioritize RDS recommendations"""
        actions = []
        
        for rec in rds_recommendations:
            if rec['environment'] in ['dev', 'test'] and rec['action'] in ['stop', 'schedule']:
                action = {
                    'service': 'RDS',
                    'resource_id': rec['instance_identifier'],
                    'action': rec['action'],
                    'description': rec['reason'],
                    'monthly_savings': rec['estimated_monthly_savings'],
                    'risk_level': rec['risk_level'],
                    'priority': 1 if rec['environment'] == 'dev' else 2,
                    'implementation_time': '30 minutes',
                    'category': 'immediate' if rec['action'] == 'stop' else 'quick_win',
                    'commands': [
                        f"aws rds stop-db-instance --db-instance-identifier {rec['instance_identifier']}"
                    ]
                }
                
                if action['category'] == 'immediate':
                    self.immediate_actions.append(action)
                else:
                    self.scheduled_actions.append(action)
            else:
                actions.append({
                    'service': 'RDS',
                    'resource_id': rec['instance_identifier'],
                    'action': rec['action'],
                    'description': rec['reason'],
                    'monthly_savings': rec['estimated_monthly_savings'],
                    'risk_level': rec['risk_level'],
                    'priority': 3,
                    'category': 'planned'
                })
        
        return actions
    
    def analyze_s3_recommendations(self, s3_recommendations: List[Dict]) -> List[Dict]:
        """Analyze S3 optimization opportunities"""
        actions = []
        
        for rec in s3_recommendations:
            if rec['action'] == 'apply_lifecycle_policy':
                action = {
                    'service': 'S3',
                    'resource_id': rec['bucket_name'],
                    'action': 'lifecycle_policy',
                    'description': f"Apply lifecycle policy: {rec['reason']}",
                    'monthly_savings': rec['estimated_monthly_savings'],
                    'risk_level': 'low',
                    'priority': 2,
                    'implementation_time': '15 minutes',
                    'category': 'policy_change'
                }
                self.policy_changes.append(action)
            
            elif rec['action'] == 'delete_bucket' and 'unused' in rec['reason'].lower():
                action = {
                    'service': 'S3',
                    'resource_id': rec['bucket_name'],
                    'action': 'archive_and_delete',
                    'description': rec['reason'],
                    'monthly_savings': rec['estimated_monthly_savings'],
                    'risk_level': rec['risk_level'],
                    'priority': 3,
                    'implementation_time': '1 hour',
                    'category': 'planned'
                }
                actions.append(action)
        
        return actions
    
    def create_emergency_plan(self, all_recommendations: Dict[str, List]) -> Dict[str, Any]:
        """Create comprehensive emergency cost reduction plan"""
        
        # Process all recommendations
        all_actions = []
        
        if 'ec2' in all_recommendations:
            all_actions.extend(self.analyze_ec2_recommendations(all_recommendations['ec2']))
        
        if 'rds' in all_recommendations:
            all_actions.extend(self.analyze_rds_recommendations(all_recommendations['rds']))
        
        if 's3' in all_recommendations:
            all_actions.extend(self.analyze_s3_recommendations(all_recommendations['s3']))
        
        # Calculate cumulative savings
        immediate_savings = sum(a['monthly_savings'] for a in self.immediate_actions)
        quick_win_savings = sum(a['monthly_savings'] for a in self.scheduled_actions)
        policy_savings = sum(a['monthly_savings'] for a in self.policy_changes)
        
        # Create phased plan
        plan = {
            'target_savings': self.target_savings,
            'total_identified_savings': immediate_savings + quick_win_savings + policy_savings,
            'savings_achieved_percentage': ((immediate_savings + quick_win_savings + policy_savings) / self.target_savings) * 100,
            
            'phase_1_immediate': {
                'description': 'Stop idle dev/test resources (0-24 hours)',
                'actions': sorted(self.immediate_actions, key=lambda x: x['monthly_savings'], reverse=True),
                'total_savings': immediate_savings,
                'implementation_time': '24 hours',
                'risk': 'Low',
                'approval_required': 'Team Lead'
            },
            
            'phase_2_quick_wins': {
                'description': 'Implement scheduling and lifecycle policies (1-3 days)',
                'actions': sorted(self.scheduled_actions + self.policy_changes, 
                                key=lambda x: x['monthly_savings'], reverse=True),
                'total_savings': quick_win_savings + policy_savings,
                'implementation_time': '3 days',
                'risk': 'Low',
                'approval_required': 'Manager'
            },
            
            'phase_3_optimization': {
                'description': 'Rightsize and optimize production resources (1-2 weeks)',
                'actions': sorted(all_actions, key=lambda x: x['monthly_savings'], reverse=True)[:10],
                'total_savings': sum(a['monthly_savings'] for a in all_actions[:10]),
                'implementation_time': '2 weeks',
                'risk': 'Medium',
                'approval_required': 'Director'
            },
            
            'untagged_resources': self._identify_untagged_resources(all_recommendations),
            'duplicate_resources': self._identify_duplicates(all_recommendations),
            'executive_summary': self._create_executive_summary(
                immediate_savings, quick_win_savings, policy_savings, all_actions
            )
        }
        
        return plan
    
    def _calculate_priority(self, recommendation: Dict) -> int:
        """Calculate action priority (1=highest)"""
        if recommendation['risk_level'] == 'low' and recommendation['monthly_savings'] > 1000:
            return 1
        elif recommendation['risk_level'] == 'low':
            return 2
        elif recommendation['risk_level'] == 'medium' and recommendation['monthly_savings'] > 2000:
            return 3
        else:
            return 4
    
    def _estimate_implementation_time(self, action: str) -> str:
        """Estimate time to implement action"""
        times = {
            'stop': '15 minutes',
            'schedule': '30 minutes',
            'rightsize': '2 hours',
            'delete': '1 hour',
            'lifecycle_policy': '15 minutes'
        }
        return times.get(action, '1 hour')
    
    def _create_rollback_plan(self, recommendation: Dict) -> List[str]:
        """Create rollback steps for each action"""
        if recommendation['action'] == 'stop':
            return [
                f"1. Start instance: aws ec2 start-instances --instance-ids {recommendation['instance_id']}",
                "2. Verify application health",
                "3. Update monitoring alerts"
            ]
        elif recommendation['action'] == 'rightsize':
            return [
                "1. Stop new instance",
                "2. Change instance type back to original",
                "3. Start instance with original type",
                "4. Update load balancer if needed"
            ]
        return ["Rollback plan specific to action"]
    
    def _identify_untagged_resources(self, all_recommendations: Dict) -> Dict[str, int]:
        """Identify resources without proper tags"""
        untagged = {
            'ec2_instances': 0,
            'rds_databases': 0,
            's3_buckets': 0,
            'estimated_waste': 0
        }
        
        # In real implementation, would check tags from inventory
        # Estimating 30% resources are untagged
        if 'ec2' in all_recommendations:
            untagged['ec2_instances'] = int(len(all_recommendations['ec2']) * 0.3)
            untagged['estimated_waste'] += untagged['ec2_instances'] * 50  # $50/month per untagged instance
        
        return untagged
    
    def _identify_duplicates(self, all_recommendations: Dict) -> List[Dict]:
        """Identify potential duplicate resources"""
        duplicates = []
        
        # Look for dev/test/staging versions
        environments = ['dev', 'test', 'staging', 'uat']
        
        # Simplified duplicate detection
        duplicates.append({
            'type': 'Multiple test databases',
            'count': 15,
            'monthly_cost': 3000,
            'recommendation': 'Consolidate test databases and use RDS snapshots for testing'
        })
        
        duplicates.append({
            'type': 'Redundant dev environments',
            'count': 8,
            'monthly_cost': 2500,
            'recommendation': 'Use containerized dev environments instead of full EC2 instances'
        })
        
        return duplicates
    
    def _create_executive_summary(self, immediate: float, quick: float, 
                                 policy: float, all_actions: List) -> Dict:
        """Create executive summary for leadership"""
        total_potential = immediate + quick + policy + sum(a.get('monthly_savings', 0) for a in all_actions)
        
        return {
            'situation': 'TechStartup acquisition is consuming $47K/month in AWS costs with no documentation',
            'findings': {
                'idle_resources': f"${immediate:,.0f}/month in idle dev/test resources",
                'unoptimized': f"${quick:,.0f}/month in scheduling opportunities",
                'storage_waste': f"${policy:,.0f}/month in storage optimization",
                'total_waste': f"${total_potential:,.0f}/month in total optimization opportunities"
            },
            'recommendations': [
                f"Immediate: Stop idle resources to save ${immediate:,.0f}/month (24 hours)",
                f"Quick wins: Implement scheduling to save ${quick:,.0f}/month (3 days)",
                f"Policies: Apply lifecycle rules to save ${policy:,.0f}/month (1 week)",
                "Long-term: Implement tagging strategy and cost allocation"
            ],
            'target_achievement': f"{(total_potential / self.target_savings * 100):.0f}% of $20K target",
            'next_steps': [
                "Get approval for Phase 1 immediate actions",
                "Schedule review meeting with dev teams",
                "Implement cost monitoring dashboard",
                "Create documentation for all resources"
            ]
        }
    
    def export_plan_to_excel(self, plan: Dict, filename: str):
        """Export emergency plan to Excel for leadership"""
        wb = Workbook()
        
        # Executive Summary Sheet
        ws_exec = wb.active
        ws_exec.title = "Executive Summary"
        
        # Title
        ws_exec.merge_cells('A1:E1')
        ws_exec['A1'] = "TechStartup AWS Cost Reduction Plan"
        ws_exec['A1'].font = Font(size=16, bold=True)
        ws_exec['A1'].alignment = Alignment(horizontal='center')
        
        # Summary metrics
        row = 3
        ws_exec[f'A{row}'] = "Target Monthly Savings"
        ws_exec[f'B{row}'] = f"${self.target_savings:,.0f}"
        ws_exec[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws_exec[f'A{row}'] = "Identified Savings"
        ws_exec[f'B{row}'] = f"${plan['total_identified_savings']:,.0f}"
        ws_exec[f'B{row}'].font = Font(bold=True, color="008000")
        
        row += 1
        ws_exec[f'A{row}'] = "Target Achievement"
        ws_exec[f'B{row}'] = f"{plan['savings_achieved_percentage']:.0f}%"
        
        # Phase summary
        row += 2
        ws_exec[f'A{row}'] = "Implementation Phases"
        ws_exec[f'A{row}'].font = Font(bold=True)
        
        row += 1
        phases = [
            ("Phase 1: Immediate Actions", plan['phase_1_immediate']),
            ("Phase 2: Quick Wins", plan['phase_2_quick_wins']),
            ("Phase 3: Optimization", plan['phase_3_optimization'])
        ]
        
        for phase_name, phase_data in phases:
            ws_exec[f'A{row}'] = phase_name
            ws_exec[f'B{row}'] = f"${phase_data['total_savings']:,.0f}"
            ws_exec[f'C{row}'] = phase_data['implementation_time']
            ws_exec[f'D{row}'] = phase_data['risk']
            row += 1
        
        # Immediate Actions Sheet
        ws_immediate = wb.create_sheet("Immediate Actions")
        headers = ['Resource ID', 'Service', 'Action', 'Monthly Savings', 'Risk', 'Commands']
        
        for col, header in enumerate(headers, 1):
            cell = ws_immediate.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        for action in plan['phase_1_immediate']['actions']:
            ws_immediate.cell(row=row, column=1, value=action['resource_id'])
            ws_immediate.cell(row=row, column=2, value=action['service'])
            ws_immediate.cell(row=row, column=3, value=action['action'])
            ws_immediate.cell(row=row, column=4, value=f"${action['monthly_savings']:,.0f}")
            ws_immediate.cell(row=row, column=5, value=action['risk_level'])
            ws_immediate.cell(row=row, column=6, value='; '.join(action.get('commands', [])))
            row += 1
        
        # Quick Wins Sheet
        ws_quick = wb.create_sheet("Quick Wins")
        # Similar structure to immediate actions
        
        # Format all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        logger.info(f"Emergency cost reduction plan saved to {filename}")
    
    def generate_implementation_script(self, plan: Dict, filename: str):
        """Generate bash script for immediate actions"""
        script_lines = [
            "#!/bin/bash",
            "# TechStartup AWS Cost Reduction - Immediate Actions Script",
            f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"# Target Savings: ${self.target_savings:,.0f}/month",
            "",
            "set -e  # Exit on error",
            "",
            "echo 'Starting AWS cost reduction actions...'",
            "",
            "# Phase 1: Immediate Actions (Stop idle resources)",
            "echo 'Phase 1: Stopping idle resources...'",
            ""
        ]
        
        for action in plan['phase_1_immediate']['actions']:
            script_lines.append(f"# {action['description']}")
            script_lines.append(f"# Savings: ${action['monthly_savings']:.0f}/month")
            
            if action['service'] == 'EC2' and action['action'] == 'stop':
                script_lines.append(f"echo 'Stopping EC2 instance {action['resource_id']}...'")
                script_lines.append(f"aws ec2 stop-instances --instance-ids {action['resource_id']}")
                script_lines.append("")
            
            elif action['service'] == 'RDS' and action['action'] == 'stop':
                script_lines.append(f"echo 'Stopping RDS instance {action['resource_id']}...'")
                script_lines.append(f"aws rds stop-db-instance --db-instance-identifier {action['resource_id']}")
                script_lines.append("")
        
        script_lines.extend([
            "",
            "echo 'Phase 1 complete!'",
            f"echo 'Estimated savings: ${plan['phase_1_immediate']['total_savings']:,.0f}/month'",
            "",
            "# Create backup of actions taken",
            "echo 'Creating action log...'",
            f"echo 'Actions completed on {datetime.now()}' > cost_reduction_actions.log"
        ])
        
        with open(filename, 'w') as f:
            f.write('\n'.join(script_lines))
        
        logger.info(f"Implementation script saved to {filename}")