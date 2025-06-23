#!/usr/bin/env python3
"""
Multi-Account AWS Resource Inventory Script
Discovers all resources across multiple AWS accounts for the TechStartup acquisition
"""
import boto3
import click
import pandas as pd
from datetime import datetime, timedelta
from typing import List, Dict, Any
import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class MultiAccountInventory:
    """Discovers and inventories resources across multiple AWS accounts"""
    
    def __init__(self, accounts: List[Dict[str, str]], regions: List[str] = None):
        self.accounts = accounts
        self.regions = regions or ['us-east-1', 'us-west-2', 'eu-west-1', 'ap-southeast-1']
        self.inventory = {
            'ec2_instances': [],
            'rds_databases': [],
            's3_buckets': [],
            'ebs_volumes': [],
            'elastic_ips': [],
            'load_balancers': [],
            'snapshots': []
        }
        self.cost_data = {}
        
    def assume_role(self, account_id: str, role_name: str) -> boto3.Session:
        """Assume role in target account"""
        sts = boto3.client('sts')
        role_arn = f"arn:aws:iam::{account_id}:role/{role_name}"
        
        try:
            response = sts.assume_role(
                RoleArn=role_arn,
                RoleSessionName=f"CostOptimizer-Inventory-{account_id}"
            )
            
            return boto3.Session(
                aws_access_key_id=response['Credentials']['AccessKeyId'],
                aws_secret_access_key=response['Credentials']['SecretAccessKey'],
                aws_session_token=response['Credentials']['SessionToken']
            )
        except Exception as e:
            logger.error(f"Failed to assume role in account {account_id}: {e}")
            raise
    
    def discover_ec2_instances(self, session: boto3.Session, account: Dict[str, str]) -> List[Dict[str, Any]]:
        """Discover all EC2 instances in an account"""
        instances = []
        
        for region in self.regions:
            try:
                ec2 = session.client('ec2', region_name=region)
                
                paginator = ec2.get_paginator('describe_instances')
                for page in paginator.paginate():
                    for reservation in page['Reservations']:
                        for instance in reservation['Instances']:
                            # Get instance details
                            instance_data = {
                                'account_id': account['account_id'],
                                'account_name': account['account_name'],
                                'region': region,
                                'instance_id': instance['InstanceId'],
                                'instance_type': instance.get('InstanceType'),
                                'state': instance['State']['Name'],
                                'launch_time': str(instance.get('LaunchTime')),
                                'availability_zone': instance.get('Placement', {}).get('AvailabilityZone'),
                                'private_ip': instance.get('PrivateIpAddress'),
                                'public_ip': instance.get('PublicIpAddress'),
                                'vpc_id': instance.get('VpcId'),
                                'subnet_id': instance.get('SubnetId'),
                                'security_groups': [sg['GroupId'] for sg in instance.get('SecurityGroups', [])],
                                'tags': {tag['Key']: tag['Value'] for tag in instance.get('Tags', [])},
                                'name': next((tag['Value'] for tag in instance.get('Tags', []) 
                                            if tag['Key'] == 'Name'), ''),
                                'environment': next((tag['Value'] for tag in instance.get('Tags', []) 
                                                   if tag['Key'] == 'Environment'), 'unknown'),
                                'monthly_cost': self._estimate_ec2_cost(instance.get('InstanceType', ''))
                            }
                            
                            instances.append(instance_data)
                            
            except Exception as e:
                logger.error(f"Error discovering EC2 in {account['account_name']}/{region}: {e}")
        
        return instances
    
    def discover_rds_databases(self, session: boto3.Session, account: Dict[str, str]) -> List[Dict[str, Any]]:
        """Discover all RDS databases in an account"""
        databases = []
        
        for region in self.regions:
            try:
                rds = session.client('rds', region_name=region)
                
                # Get RDS instances
                paginator = rds.get_paginator('describe_db_instances')
                for page in paginator.paginate():
                    for db in page['DBInstances']:
                        db_data = {
                            'account_id': account['account_id'],
                            'account_name': account['account_name'],
                            'region': region,
                            'db_instance_identifier': db['DBInstanceIdentifier'],
                            'db_instance_class': db['DBInstanceClass'],
                            'engine': db['Engine'],
                            'engine_version': db.get('EngineVersion'),
                            'status': db['DBInstanceStatus'],
                            'allocated_storage': db['AllocatedStorage'],
                            'storage_type': db.get('StorageType'),
                            'multi_az': db.get('MultiAZ', False),
                            'backup_retention': db.get('BackupRetentionPeriod'),
                            'creation_time': str(db.get('InstanceCreateTime')),
                            'endpoint': db.get('Endpoint', {}).get('Address'),
                            'vpc_id': db.get('DBSubnetGroup', {}).get('VpcId'),
                            'tags': {tag['Key']: tag['Value'] for tag in db.get('TagList', [])},
                            'environment': next((tag['Value'] for tag in db.get('TagList', []) 
                                               if tag['Key'] == 'Environment'), 'unknown'),
                            'monthly_cost': self._estimate_rds_cost(db['DBInstanceClass'], db.get('MultiAZ', False))
                        }
                        
                        databases.append(db_data)
                
                # Get Aurora clusters
                try:
                    cluster_paginator = rds.get_paginator('describe_db_clusters')
                    for page in cluster_paginator.paginate():
                        for cluster in page['DBClusters']:
                            cluster_data = {
                                'account_id': account['account_id'],
                                'account_name': account['account_name'],
                                'region': region,
                                'db_cluster_identifier': cluster['DBClusterIdentifier'],
                                'engine': cluster['Engine'],
                                'status': cluster['Status'],
                                'cluster_members': len(cluster.get('DBClusterMembers', [])),
                                'allocated_storage': cluster.get('AllocatedStorage', 0),
                                'backup_retention': cluster.get('BackupRetentionPeriod'),
                                'creation_time': str(cluster.get('ClusterCreateTime')),
                                'is_cluster': True,
                                'monthly_cost': self._estimate_aurora_cluster_cost(cluster)
                            }
                            databases.append(cluster_data)
                except:
                    pass  # Aurora might not be available in all regions
                    
            except Exception as e:
                logger.error(f"Error discovering RDS in {account['account_name']}/{region}: {e}")
        
        return databases
    
    def discover_s3_buckets(self, session: boto3.Session, account: Dict[str, str]) -> List[Dict[str, Any]]:
        """Discover all S3 buckets and calculate total storage"""
        buckets = []
        s3 = session.client('s3')
        cloudwatch = session.client('cloudwatch')
        
        try:
            # List all buckets
            response = s3.list_buckets()
            
            for bucket in response['Buckets']:
                bucket_name = bucket['Name']
                
                try:
                    # Get bucket location
                    location = s3.get_bucket_location(Bucket=bucket_name)
                    region = location.get('LocationConstraint') or 'us-east-1'
                    
                    # Get bucket size from CloudWatch
                    size_bytes = self._get_bucket_size(cloudwatch, bucket_name)
                    
                    # Get bucket tags
                    try:
                        tag_response = s3.get_bucket_tagging(Bucket=bucket_name)
                        tags = {tag['Key']: tag['Value'] for tag in tag_response.get('TagSet', [])}
                    except:
                        tags = {}
                    
                    # Get versioning status
                    try:
                        versioning = s3.get_bucket_versioning(Bucket=bucket_name)
                        versioning_status = versioning.get('Status', 'Disabled')
                    except:
                        versioning_status = 'Unknown'
                    
                    # Get lifecycle configuration
                    try:
                        lifecycle = s3.get_bucket_lifecycle_configuration(Bucket=bucket_name)
                        has_lifecycle = len(lifecycle.get('Rules', [])) > 0
                    except:
                        has_lifecycle = False
                    
                    bucket_data = {
                        'account_id': account['account_id'],
                        'account_name': account['account_name'],
                        'bucket_name': bucket_name,
                        'region': region,
                        'creation_date': str(bucket['CreationDate']),
                        'size_bytes': size_bytes,
                        'size_gb': size_bytes / (1024**3),
                        'size_tb': size_bytes / (1024**4),
                        'object_count': self._get_object_count(s3, bucket_name),
                        'versioning': versioning_status,
                        'has_lifecycle': has_lifecycle,
                        'tags': tags,
                        'environment': tags.get('Environment', 'unknown'),
                        'monthly_cost': self._estimate_s3_cost(size_bytes)
                    }
                    
                    buckets.append(bucket_data)
                    
                except Exception as e:
                    logger.error(f"Error processing bucket {bucket_name}: {e}")
                    
        except Exception as e:
            logger.error(f"Error discovering S3 in {account['account_name']}: {e}")
        
        return buckets
    
    def _get_bucket_size(self, cloudwatch: Any, bucket_name: str) -> int:
        """Get bucket size from CloudWatch metrics"""
        try:
            response = cloudwatch.get_metric_statistics(
                Namespace='AWS/S3',
                MetricName='BucketSizeBytes',
                Dimensions=[
                    {'Name': 'BucketName', 'Value': bucket_name},
                    {'Name': 'StorageType', 'Value': 'StandardStorage'}
                ],
                StartTime=datetime.utcnow() - timedelta(days=2),
                EndTime=datetime.utcnow(),
                Period=86400,
                Statistics=['Average']
            )
            
            if response['Datapoints']:
                return int(response['Datapoints'][-1]['Average'])
        except:
            pass
        
        return 0
    
    def _get_object_count(self, s3: Any, bucket_name: str) -> int:
        """Get approximate object count"""
        try:
            paginator = s3.get_paginator('list_objects_v2')
            count = 0
            
            # Just get first page for estimate
            for page in paginator.paginate(Bucket=bucket_name, MaxKeys=1000):
                count += len(page.get('Contents', []))
                if count >= 1000:  # Stop after first page for performance
                    return f"{count}+"
                    
            return count
        except:
            return 0
    
    def discover_all_resources(self) -> Dict[str, Any]:
        """Discover all resources across all accounts"""
        logger.info(f"Starting discovery across {len(self.accounts)} accounts")
        
        for account in self.accounts:
            logger.info(f"Processing account: {account['account_name']} ({account['account_id']})")
            
            try:
                # Assume role in the account
                session = self.assume_role(account['account_id'], account['role_name'])
                
                # Discover resources in parallel
                with ThreadPoolExecutor(max_workers=3) as executor:
                    futures = {
                        executor.submit(self.discover_ec2_instances, session, account): 'ec2',
                        executor.submit(self.discover_rds_databases, session, account): 'rds',
                        executor.submit(self.discover_s3_buckets, session, account): 's3'
                    }
                    
                    for future in as_completed(futures):
                        resource_type = futures[future]
                        try:
                            resources = future.result()
                            
                            if resource_type == 'ec2':
                                self.inventory['ec2_instances'].extend(resources)
                            elif resource_type == 'rds':
                                self.inventory['rds_databases'].extend(resources)
                            elif resource_type == 's3':
                                self.inventory['s3_buckets'].extend(resources)
                                
                            logger.info(f"Found {len(resources)} {resource_type} resources in {account['account_name']}")
                            
                        except Exception as e:
                            logger.error(f"Error discovering {resource_type} in {account['account_name']}: {e}")
                            
            except Exception as e:
                logger.error(f"Error processing account {account['account_name']}: {e}")
        
        return self.inventory
    
    def generate_summary_report(self) -> Dict[str, Any]:
        """Generate summary statistics matching the practice problem"""
        summary = {
            'total_accounts': len(self.accounts),
            'total_regions': len(self.regions),
            'discovery_timestamp': datetime.utcnow().isoformat(),
            
            'ec2_summary': {
                'total_instances': len(self.inventory['ec2_instances']),
                'running_instances': len([i for i in self.inventory['ec2_instances'] if i['state'] == 'running']),
                'stopped_instances': len([i for i in self.inventory['ec2_instances'] if i['state'] == 'stopped']),
                'by_type': {},
                'by_environment': {},
                'total_monthly_cost': sum(i['monthly_cost'] for i in self.inventory['ec2_instances'])
            },
            
            'rds_summary': {
                'total_databases': len(self.inventory['rds_databases']),
                'by_engine': {},
                'by_environment': {},
                'multi_az_count': len([d for d in self.inventory['rds_databases'] if d.get('multi_az')]),
                'total_monthly_cost': sum(d['monthly_cost'] for d in self.inventory['rds_databases'])
            },
            
            's3_summary': {
                'total_buckets': len(self.inventory['s3_buckets']),
                'total_storage_tb': sum(b['size_tb'] for b in self.inventory['s3_buckets']),
                'total_objects': sum(b['object_count'] if isinstance(b['object_count'], int) else 1000 
                                   for b in self.inventory['s3_buckets']),
                'buckets_without_lifecycle': len([b for b in self.inventory['s3_buckets'] if not b['has_lifecycle']]),
                'total_monthly_cost': sum(b['monthly_cost'] for b in self.inventory['s3_buckets'])
            },
            
            'cost_summary': {
                'total_monthly_cost': 0,
                'total_annual_cost': 0,
                'by_service': {},
                'by_account': {},
                'by_environment': {}
            }
        }
        
        # Calculate detailed summaries
        for instance in self.inventory['ec2_instances']:
            instance_type = instance['instance_type']
            environment = instance['environment']
            
            summary['ec2_summary']['by_type'][instance_type] = \
                summary['ec2_summary']['by_type'].get(instance_type, 0) + 1
            summary['ec2_summary']['by_environment'][environment] = \
                summary['ec2_summary']['by_environment'].get(environment, 0) + 1
        
        for db in self.inventory['rds_databases']:
            engine = db['engine']
            environment = db['environment']
            
            summary['rds_summary']['by_engine'][engine] = \
                summary['rds_summary']['by_engine'].get(engine, 0) + 1
            summary['rds_summary']['by_environment'][environment] = \
                summary['rds_summary']['by_environment'].get(environment, 0) + 1
        
        # Calculate total costs
        total_monthly = (summary['ec2_summary']['total_monthly_cost'] +
                        summary['rds_summary']['total_monthly_cost'] +
                        summary['s3_summary']['total_monthly_cost'])
        
        summary['cost_summary']['total_monthly_cost'] = total_monthly
        summary['cost_summary']['total_annual_cost'] = total_monthly * 12
        
        summary['cost_summary']['by_service'] = {
            'EC2': summary['ec2_summary']['total_monthly_cost'],
            'RDS': summary['rds_summary']['total_monthly_cost'],
            'S3': summary['s3_summary']['total_monthly_cost']
        }
        
        # Alert on key metrics from practice problem
        logger.info("\n" + "="*50)
        logger.info("TECHSTARTUP ACQUISITION - AWS INVENTORY SUMMARY")
        logger.info("="*50)
        logger.info(f"Total EC2 Instances: {summary['ec2_summary']['total_instances']} (Target: 147)")
        logger.info(f"Total RDS Databases: {summary['rds_summary']['total_databases']} (Target: 37)")
        logger.info(f"Total S3 Storage: {summary['s3_summary']['total_storage_tb']:.2f} TB (Target: 890 TB)")
        logger.info(f"Total Monthly Cost: ${summary['cost_summary']['total_monthly_cost']:,.2f} (Target: ~$47,000)")
        logger.info("="*50 + "\n")
        
        return summary
    
    def export_to_excel(self, filename: str):
        """Export inventory to Excel with multiple sheets"""
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Summary sheet
            summary = self.generate_summary_report()
            summary_df = pd.DataFrame([
                {'Metric': 'Total AWS Accounts', 'Value': summary['total_accounts']},
                {'Metric': 'Total EC2 Instances', 'Value': summary['ec2_summary']['total_instances']},
                {'Metric': 'Total RDS Databases', 'Value': summary['rds_summary']['total_databases']},
                {'Metric': 'Total S3 Buckets', 'Value': summary['s3_summary']['total_buckets']},
                {'Metric': 'Total S3 Storage (TB)', 'Value': f"{summary['s3_summary']['total_storage_tb']:.2f}"},
                {'Metric': 'Total Monthly Cost', 'Value': f"${summary['cost_summary']['total_monthly_cost']:,.2f}"},
                {'Metric': 'Total Annual Cost', 'Value': f"${summary['cost_summary']['total_annual_cost']:,.2f}"}
            ])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # EC2 instances sheet
            if self.inventory['ec2_instances']:
                ec2_df = pd.DataFrame(self.inventory['ec2_instances'])
                ec2_df = ec2_df.sort_values(['account_name', 'region', 'instance_id'])
                ec2_df.to_excel(writer, sheet_name='EC2_Instances', index=False)
            
            # RDS databases sheet
            if self.inventory['rds_databases']:
                rds_df = pd.DataFrame(self.inventory['rds_databases'])
                rds_df = rds_df.sort_values(['account_name', 'region'])
                rds_df.to_excel(writer, sheet_name='RDS_Databases', index=False)
            
            # S3 buckets sheet
            if self.inventory['s3_buckets']:
                s3_df = pd.DataFrame(self.inventory['s3_buckets'])
                s3_df = s3_df.sort_values('size_tb', ascending=False)
                s3_df.to_excel(writer, sheet_name='S3_Buckets', index=False)
            
            # Format the Excel file
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format headers
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
        
        logger.info(f"Exported inventory to {filename}")
    
    def _estimate_ec2_cost(self, instance_type: str) -> float:
        """Estimate monthly cost for EC2 instance"""
        # Simplified pricing - in production use AWS Pricing API
        hourly_rates = {
            't2.micro': 0.0116, 't2.small': 0.023, 't2.medium': 0.0464,
            't3.micro': 0.0104, 't3.small': 0.0208, 't3.medium': 0.0416, 't3.large': 0.0832,
            'm5.large': 0.096, 'm5.xlarge': 0.192, 'm5.2xlarge': 0.384,
            'c5.large': 0.085, 'c5.xlarge': 0.17
        }
        
        hourly_rate = hourly_rates.get(instance_type, 0.10)  # Default rate
        return hourly_rate * 24 * 30
    
    def _estimate_rds_cost(self, instance_class: str, multi_az: bool) -> float:
        """Estimate monthly cost for RDS instance"""
        hourly_rates = {
            'db.t3.micro': 0.017, 'db.t3.small': 0.034, 'db.t3.medium': 0.068,
            'db.m5.large': 0.171, 'db.m5.xlarge': 0.342, 'db.m5.2xlarge': 0.684,
            'db.r5.large': 0.24, 'db.r5.xlarge': 0.48
        }
        
        hourly_rate = hourly_rates.get(instance_class, 0.20)
        if multi_az:
            hourly_rate *= 2
            
        return hourly_rate * 24 * 30
    
    def _estimate_aurora_cluster_cost(self, cluster: Dict[str, Any]) -> float:
        """Estimate monthly cost for Aurora cluster"""
        # Simplified - $0.10 per ACU-hour
        return 0.10 * 24 * 30 * len(cluster.get('DBClusterMembers', []))
    
    def _estimate_s3_cost(self, size_bytes: int) -> float:
        """Estimate monthly S3 storage cost"""
        size_gb = size_bytes / (1024**3)
        
        # Tiered pricing
        if size_gb <= 50000:  # First 50 TB
            return size_gb * 0.023
        elif size_gb <= 450000:  # Next 450 TB
            return 50000 * 0.023 + (size_gb - 50000) * 0.022
        else:
            return 50000 * 0.023 + 400000 * 0.022 + (size_gb - 450000) * 0.021